"""
Synthetic Phenotypic Trait Dataset Generator — Version 2
Institute of Plant Breeding, UPLB — Cereal Crops Section

Generates 500 records that faithfully replicate:
  • Exact column names and order from Basic_Information_May_2023_Autosaved.xlsx
  • Per-column fill rates (null frequencies)
  • Categorical distributions and value formats (including case/spelling variants)
  • Numeric ranges, means, and standard deviations
  • All inter-column logical constraints listed below

Logical constraints enforced
─────────────────────────────
 1. Region → Province → Town/Municipality → Dialect (hierarchical geography)
 2. Kernel Type → Kernel Color (col 3, raw) probability tables
 3. Kernel Color → ALEURONE COLOR (col 64) + ENDOSPERM COLOR (col 65)
 4. Kernel Color → Kernel color (col 84, standardised Title-Case) + Cob color (col 85)
 5. Kernel Type → Type of grain in middle third (col 82): flint→Flint, glutinous→Waxy, dent→Dent
 6. Purple-pigmentation pleiotropy: a latent pigment score derived from kernel color
    simultaneously drives Silk color (2015), Anther color (2015), Stem color,
    Leaf sheath color, Spikelet base, Palea/lemma coloration
 7. Plant Height (cm) → Ear Height (cm): EH ≈ 0.57 × PH ± noise, always EH < PH
 8. Time of anthesis (DAP) → Time of silk Emergence (DAP): silking ≥ anthesis
 9. Ear shape (col 80) → Grain row arrangement (col 81): conical ears → irregular rows
10. Downy Mildew (col 129) → Downy Mildew Value (col 128): value derived from rating
11. Bacterial Stalk Rot (col 130) → Bacterial Stalk Rot Value (col 131)
12. Waterlogging (col 119) + waterlogging trial data (cols 120-124) are co-screened
13. Bt all in data (col 156) + GT all in (col 157) → BtGt Positive/Remarks (col 159)
14. APN2 (col 155) mirrors APN (col 0) for screened records
15. Topography ↔ Drainage ↔ Irrigation co-vary realistically
16. Collecting Source → Genetic Status (farmland collection → NATIVE/LANDRACE)
17. Amylose + Amylopectin ≈ Starch (±5%)
18. Crude fat Rank (col 100) and DPPH rank (col 113) are rank positions within the batch
19. "Cereals' Breeding" note matches expected seed regeneration activities
20. Corn Borer Remarks (col 135) is derived from Corn Borer (LDA) (col 133)
"""

import random
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)
np.random.seed(42)

N = 500  # records to generate

# ══════════════════════════════════════════════════════════════════
# 1. HELPER UTILITIES
# ══════════════════════════════════════════════════════════════════


def wc(population, weights):
    """Weighted choice from a list."""
    return random.choices(population, weights=weights, k=1)[0]


def maybe(prob, value, fallback=None):
    """Return value with probability prob, else fallback."""
    return value if random.random() < prob else fallback


def nd(mu, sigma, lo, hi, decimals=1):
    """Clipped normal deviate."""
    v = np.random.normal(mu, sigma)
    return round(float(np.clip(v, lo, hi)), decimals)


# ══════════════════════════════════════════════════════════════════
# 2. GEOGRAPHY LOOKUP TABLES (derived from source data)
# ══════════════════════════════════════════════════════════════════

REGION_DATA = {
    "VI": {
        "w": 682,
        "provinces": [
            ("Antique", 215),
            ("Iloilo", 198),
            ("Capiz", 108),
            ("Aklan", 80),
            ("Negros Occidental", 55),
        ],
        "dialects": ["AKLANON", "HILIGAYNON", "KINARAY-A", "ILONGGO", "VISAYAN"],
    },
    "VII": {
        "w": 345,
        "provinces": [
            ("Cebu", 188),
            ("Bohol", 29),
            ("Negros Oriental", 29),
            ("Siquijor", 24),
        ],
        "dialects": ["CEBUANO", "VISAYAN", "BISAYA"],
    },
    "V": {
        "w": 338,
        "provinces": [("Masbate", 175), ("Camarines Sur", 138), ("Albay", 24)],
        "dialects": ["BICOL", "MASBATEÑO", "Bisaya"],
    },
    "IVA": {
        "w": 308,
        "provinces": [
            ("Batangas", 194),
            ("Quezon", 69),
            ("Cavite", 19),
            ("Laguna", 10),
        ],
        "dialects": ["TAGALOG", "Tagalog", "Filipino"],
    },
    "CARAGA": {
        "w": 286,
        "provinces": [
            ("Agusan Del Sur", 80),
            ("Surigao Del Sur", 77),
            ("Agusan Del Norte", 26),
            ("Surigao Del Norte", 21),
        ],
        "dialects": ["BISAYA", "Bisaya", "Manobo", "Cebuano"],
    },
    "IX": {
        "w": 251,
        "provinces": [
            ("Zamboanga Del Norte", 160),
            ("Zamboanga Sibugay", 33),
            ("Zamboanga Del Sur", 29),
            ("Zamboanga City", 20),
        ],
        "dialects": ["SUBANEN", "VISAYAN", "CEBUANO", "CHAVACANO"],
    },
    "IVB": {
        "w": 208,
        "provinces": [
            ("Romblon", 63),
            ("Occidental Mindoro", 56),
            ("Marinduque", 36),
            ("Oriental Mindoro", 26),
            ("Palawan", 24),
        ],
        "dialects": ["TAGALOG", "BISAYA", "Tagalog"],
    },
    "X": {
        "w": 199,
        "provinces": [
            ("Bukidnon", 108),
            ("Misamis Occidental", 42),
            ("Misamis Oriental", 13),
        ],
        "dialects": ["CEBUANO", "BISAYA", "Binisaya", "Cebuano"],
    },
    "VIII": {
        "w": 184,
        "provinces": [("Leyte", 158), ("Samar", 6)],
        "dialects": ["Cebuano", "CEBUANO", "WARAYWARAY", "Bisaya"],
    },
    "XII": {
        "w": 175,
        "provinces": [
            ("North Cotabato", 61),
            ("South Cotabato", 37),
            ("Sarangani", 23),
            ("Sultan Kudarat", 19),
        ],
        "dialects": ["BISAYA", "ILONGGO", "T'BOLI", "MAGUINDANAON"],
    },
    "II": {
        "w": 168,
        "provinces": [("Isabela", 83), ("Cagayan", 54), ("Batanes", 5), ("Quirino", 3)],
        "dialects": ["IBANAG, ILOCANO", "ILOCANO", "ITAWIS, IBANAG, ILOCANO"],
    },
    "I": {
        "w": 122,
        "provinces": [
            ("Ilocos Sur", 37),
            ("Pangasinan", 37),
            ("La Union", 32),
            ("Ilocos Norte", 15),
        ],
        "dialects": ["ILOCO", "ILOKANO", "ILOCANO", "PANGASINAN"],
    },
    "CAR": {
        "w": 105,
        "provinces": [
            ("Benguet", 34),
            ("Mt. Province", 30),
            ("Abra", 16),
            ("Kalinga", 10),
        ],
        "dialects": ["KANKANA-EY", "KAN-KANAEY", "Kankanaey", "IBALOI"],
    },
    "XI": {
        "w": 104,
        "provinces": [
            ("Davao Oriental", 22),
            ("Compostela Valley", 11),
            ("Davao Del Sur", 11),
            ("Davao Del Norte", 7),
        ],
        "dialects": ["VISAYAN", "Bisaya", "CEBUANO"],
    },
    "BARMM": {
        "w": 76,
        "provinces": [("Maguindanao", 41), ("Lanao Del Sur", 12)],
        "dialects": ["MAGUINDANAON", "MARANAO", "BISAYA"],
    },
}

PROVINCE_MUNICIPALITIES = {
    "Antique": [
        "Tobias Fornier",
        "San Remegio",
        "Hamtic",
        "Bugasong",
        "Valderama",
        "Barbaza",
        "Patnongon",
    ],
    "Iloilo": [
        "Alimodian",
        "Igbaras",
        "Maasin",
        "Guimbal",
        "Tubungan",
        "Jaro",
        "Pavia",
        "Leon",
    ],
    "Capiz": ["Jamindan", "Tapaz", "Sigma", "Mambusao", "Dumalag", "Cuartero"],
    "Aklan": ["Lezo", "Banga", "Kalibo", "Libacao", "Madalag", "Makato", "Numancia"],
    "Negros Occidental": ["Kabankalan", "Cauayan", "Silay", "La Castellana", "Murcia"],
    "Cebu": [
        "Bantayan",
        "Daanbantayan",
        "Bogo",
        "San Francisco",
        "Poro, Camotes",
        "Tuburan",
        "Medellin",
    ],
    "Bohol": ["Alicia", "Sagbayan", "San Miguel", "Loay", "Inabanga"],
    "Negros Oriental": ["Guihulngan", "Canlaon", "Bindoy", "Pamplona"],
    "Siquijor": ["Lazi", "Larena", "San Juan", "Maria"],
    "Masbate": [
        "Cataingan",
        "Monreal",
        "San Pascual",
        "Pio V. Corpuz",
        "Dimalasang",
        "Mandaon",
        "Cawayan",
    ],
    "Camarines Sur": [
        "Calabanga",
        "Buhi",
        "Iriga City",
        "Baao",
        "Nabua",
        "Libmanan",
        "Gainza",
    ],
    "Albay": ["Ligao City", "Oas", "Polangui", "Guinobatan", "Jovellar"],
    "Batangas": [
        "Tuy",
        "Lipa City",
        "Batangas City",
        "Laurel",
        "Calaca",
        "Balayan",
        "Cuenca",
        "Malvar",
    ],
    "Quezon": ["Catanauan", "Gumaca", "Macalelon", "Mulanay", "San Andres"],
    "Agusan Del Sur": ["Bunawan", "Loreto", "Talacogon", "San Luis", "Bayugan"],
    "Surigao Del Sur": ["Tagbina", "San Francisco", "Cantilan", "Hinatuan", "Madrid"],
    "Agusan Del Norte": ["Buenavista", "Cabadbaran", "Santiago", "Carmen", "Jabonga"],
    "Zamboanga Del Norte": [
        "Sindangan",
        "Leon Postigo",
        "Liloy",
        "Gutalac",
        "Tampilisan",
        "Sirawai",
    ],
    "Zamboanga Del Sur": ["Dumingag", "Tukuran", "Labangan", "Molave", "Pagadian"],
    "Romblon": ["Cajidiocan", "Odiongan", "Santa Fe", "Magdiwang", "San Fernando"],
    "Occidental Mindoro": [
        "Sablayan",
        "San Jose",
        "Abra de Ilog",
        "Calintaan",
        "Magsaysay",
    ],
    "Marinduque": ["Gasan", "Buenavista", "Torrijos", "Santa Cruz"],
    "Bukidnon": [
        "Malaybalay",
        "Valencia",
        "Sumilao",
        "Lantapan",
        "Impasug-Ong",
        "Maramag",
    ],
    "Misamis Occidental": ["Tangub City", "Oroquieta", "Clarin", "Calamba", "Plaridel"],
    "Leyte": ["Mahaplag", "Baybay City", "Abuyog", "Inopacan", "Tanauan", "Ormoc City"],
    "North Cotabato": ["Aleosan", "Carmen", "Kabacan", "Midsayap", "Pigkawayan"],
    "South Cotabato": ["Koronadal", "Surallah", "Banga", "Tupi", "Lake Sebu"],
    "Isabela": [
        "Roxas",
        "Benito Soliven",
        "Delfin Albano",
        "Alicia",
        "Jones",
        "Cauayan",
    ],
    "Cagayan": ["Peñablanca", "Sto. Niño", "Aparri", "Enrile", "Solana"],
    "Ilocos Sur": ["Narvacan", "Santa Maria", "Tagudin", "Candon City", "Vigan City"],
    "Pangasinan": ["Alcala", "Mangatarem", "San Manuel", "Urbiztondo", "Bugallon"],
    "La Union": ["Burgos", "Bacnotan", "Balaoan", "Bangar", "Caba"],
    "Benguet": ["La Trinidad", "Buguias", "Kabayan", "Kibungan", "Atok"],
    "Mt. Province": ["Bontoc", "Bauko", "Barlig", "Besao", "Natonin"],
    "Abra": ["Bangued", "Lagayan", "Malibcong", "Peñarrubia", "Tineg"],
    "Davao Oriental": ["Baganga", "Caraga", "Cateel", "Governor Generoso", "Mati"],
    "Maguindanao": ["Datu Piang", "Datu Blah Sinsuat", "Talayan", "Buluan"],
    "Lanao Del Sur": ["Saguiaran", "Malabang", "Balabagan", "Wao", "Tubaran"],
}

BARANGAYS = [
    "Poblacion",
    "Tiogan",
    "Balactasan",
    "Agcagay",
    "Sto. Niño",
    "San Juan",
    "Victoria",
    "Estampar",
    "Sta. Teresita",
    "Ticad",
    "Limpapa",
    "Manan-Ao",
    "Matab-ang",
    "Igpalge",
    "Punta",
    "Sta. Cruz",
    "Labog",
    "Salawag",
    "Cagdianao",
    "Haguimit",
    "Naga-Naga",
    "Cabangahan",
    "Lanas",
    "Cambanogoy",
    "Paitan Norte",
    "Paitan Sur",
    "Proper",
    "Zone 1",
    "Zone 2",
    "Zone 3",
    "Purok 1",
    "Purok 2",
    "Purok 3",
    "Purok 4",
    "Badlan",
    "Centro",
    "Bunga",
    "Tinago",
    "Bagong Silang",
    "New Poblacion",
    "Old Poblacion",
    "Rizal",
    "Mabini",
    "Magallanes",
    "Bonifacio",
    "Del Pilar",
    "Quezon",
    "Laurel",
    "Aguinaldo",
    "Luna",
    "Burgos",
    "Heneral",
    "Magsaysay",
    "Rosal",
    "Sampaguita",
    "Ilang-ilang",
    "Camia",
    "Cadena de Amor",
    "Dampalan",
    "Dapdap",
    "Dayakan",
    "Lugas",
    "Mabuhay",
    "Mabulo",
    "Malbago",
    "Malcampo",
    "Managok",
    "Mantalongon",
    "Mapula",
    "Maslog",
    "Matin-ao",
    "Sudlon",
    "Sunog",
    "Tabunok",
    "Tagbuane",
    "Tahup",
    "Tanon",
    "Tapilon",
    "Tarong",
    "Tubigan",
    "Tubod",
    "Tugas",
    "Tulay",
    "Tunga",
    "Tungkop",
]

SITIOS = [
    "-",
    "Proper",
    "Purok 1",
    "Purok 2",
    "Zone 2",
    "Zone 3",
    "Manan-Ao",
    "Badlan",
    "Centro",
    "P-2",
    "P-3",
    "Purok 4",
    "Purok 3",
    "Upper",
    "Lower",
    "Crossing",
    "Sitio Mabini",
    "Sitio Rizal",
    "Sitio Uno",
    "Sitio Dos",
    "Sitio Tres",
    "Sitio Cuatro",
    "Ilaya",
    "Ibaba",
    "Ilog",
    "Talon",
    "Batis",
    "Bukid",
    "Bundok",
]


# ══════════════════════════════════════════════════════════════════
# 3. KERNEL GENETICS
# ══════════════════════════════════════════════════════════════════

KERNEL_TYPES_W = [
    ("flint", 2156),
    ("glutinous", 755),
    ("Flint", 425),
    ("Flint/glutinous", 106),
    ("Glutinous", 61),
    ("flint/glutinous", 26),
    ("dent", 23),
    ("flint/dent", 19),
]


def _norm_kt(kt):
    k = kt.lower()
    if "glutin" in k and "flint" in k:
        return "flint/glutinous"
    if "glutin" in k:
        return "glutinous"
    if "dent" in k and "flint" in k:
        return "flint/dent"
    if "dent" in k:
        return "dent"
    return "flint"


KCOLOR_BY_KT = {
    "flint": [
        ("white", 0.535),
        ("White", 0.165),
        ("yellow", 0.085),
        ("yellow/orange", 0.030),
        ("yellow/red", 0.030),
        ("white/red", 0.050),
        ("white/redish", 0.025),
        ("white/reddish white", 0.020),
        ("white/purple", 0.025),
        ("red", 0.018),
        ("orange", 0.012),
        ("purple", 0.005),
    ],
    "glutinous": [
        ("white", 0.720),
        ("White", 0.200),
        ("white/purple", 0.040),
        ("white/red", 0.020),
        ("white/redish", 0.010),
        ("purple", 0.005),
        ("red", 0.005),
    ],
    "flint/glutinous": [
        ("white", 0.600),
        ("White", 0.200),
        ("white/red", 0.060),
        ("white/purple", 0.050),
        ("yellow", 0.040),
        ("white/redish", 0.030),
        ("red", 0.020),
    ],
    "dent": [
        ("white", 0.500),
        ("White", 0.200),
        ("yellow", 0.160),
        ("red", 0.060),
        ("orange", 0.040),
        ("purple", 0.040),
    ],
    "flint/dent": [
        ("white", 0.500),
        ("White", 0.185),
        ("yellow", 0.150),
        ("red", 0.055),
        ("orange", 0.055),
        ("purple", 0.055),
    ],
}


def gen_kernel_color(kernel_type):
    nt = _norm_kt(kernel_type)
    pool = KCOLOR_BY_KT.get(nt, KCOLOR_BY_KT["flint"])
    vals, ws_ = zip(*pool)
    return wc(vals, ws_)


def get_pigment(kc):
    k = kc.lower()
    if "purple" in k:
        return random.uniform(0.65, 1.0)
    if k.startswith("red") or "red" in k:
        return random.uniform(0.35, 0.75)
    if "orange" in k:
        return random.uniform(0.10, 0.35)
    if "yellow" in k:
        return random.uniform(0.00, 0.15)
    return random.uniform(0.00, 0.20)  # white


def gen_aleurone(kc):
    k = kc.lower()
    if "purple" in k:
        return wc(
            [
                "colorless/purple",
                "colorless/red/purple",
                "colorless/purple/red",
                "purple",
            ],
            [60, 25, 10, 5],
        )
    if "red" in k:
        return wc(
            ["colorless/red", "red", "colorless/red/purple", "colorless"],
            [60, 20, 15, 5],
        )
    if "orange" in k:
        return wc(["colorless/orange", "colorless", "colorless/red"], [60, 30, 10])
    if "yellow" in k:
        return wc(["colorless", "colorless/red"], [90, 10])
    return "colorless"  # white


def gen_endosperm(kc, kt):
    k = kc.lower()
    nt = _norm_kt(kt)
    if "glutin" in nt:
        return "white"
    if "yellow" in k or "orange" in k:
        return wc(["yellow", "white/yellow", "white"], [75, 15, 10])
    return "white"


def gen_kernel_color_std(kc):
    """Standardised Title-Case kernel color for col 84."""
    k = kc.lower()
    if "yellow" in k and "orange" in k:
        return wc(["Yellow orange", "Yellowish white"], [60, 40])
    if "yellow" in k:
        return wc(["Yellowish white", "Yellow orange"], [70, 30])
    if "purple" in k:
        return "Purple"
    if "orange" in k and "red" in k:
        return "Red orange"
    if "orange" in k:
        return "Orange"
    if "red" in k:
        return wc(["Red", "Dark red"], [80, 20])
    return "White"


def gen_cob_color(kc):
    k = kc.lower()
    if "red" in k:
        return wc(["White", "Red", "White , Red", "White,Red"], [30, 20, 30, 20])
    if "purple" in k:
        return wc(["White", "White,Purple", "White , Purple"], [40, 35, 25])
    if "orange" in k:
        return wc(["White", "White , Red"], [85, 15])
    if "yellow" in k:
        return wc(["White", "White , Red"], [90, 10])
    return "White"


def gen_grain_type(kt):
    nt = _norm_kt(kt)
    if nt == "glutinous":
        return wc(["Waxy", "Flint", "Opaque"], [80, 12, 8])
    if nt == "flint":
        return wc(["Flint", "Waxy", "Semi-flint", "Opaque"], [75, 10, 10, 5])
    if nt == "dent":
        return wc(["Dent", "Semi-dent", "Flint"], [50, 30, 20])
    if nt == "flint/glutinous":
        return wc(["Flint", "Waxy", "Semi-flint"], [45, 40, 15])
    if nt == "flint/dent":
        return wc(["Flint", "Dent", "Semi-dent", "Semi-flint"], [40, 30, 20, 10])
    return "Flint"


# ══════════════════════════════════════════════════════════════════
# 4. PIGMENT-DRIVEN COLOUR TRAITS (2015 MEASUREMENTS)
# ══════════════════════════════════════════════════════════════════


def gen_silk_color_2015(pig):
    if pig > 0.6:
        return wc(["purple", "purplish", "purplish/pale green"], [70, 20, 10])
    if pig > 0.3:
        return wc(["purplish", "purplish/pale green", "purple"], [50, 30, 20])
    return wc(["purplish/pale green", "purplish"], [60, 40])


def gen_anther_2015(pig):
    if pig > 0.6:
        return wc(["purple", "purple/purplish", "purplish"], [55, 25, 20])
    if pig > 0.3:
        return wc(["purple/yellow", "purple", "purplish/yellow"], [40, 35, 25])
    return wc(["purple/yellow", "purplish/yellow", "yellow"], [40, 35, 25])


def gen_stem_color_2015(pig):
    if pig > 0.5:
        return wc(["purplish", "purple/green", "green/purplish"], [40, 35, 25])
    if pig > 0.2:
        return wc(["green/purplish", "purplish", "pale green/green"], [50, 25, 25])
    return wc(["green", "pale green/green", "green/purplish"], [70, 20, 10])


def gen_sheath_2015(pig):
    if pig > 0.5:
        return wc(["purplish", "purple/green", "green/purplish"], [40, 35, 25])
    return wc(["green", "green/purplish", "purplish"], [80, 15, 5])


def gen_spikelet_2015(pig):
    if pig > 0.5:
        return wc(["purple/greenish", "purple", "greenish"], [50, 30, 20])
    return wc(["greenish", "purple/greenish"], [80, 20])


def gen_palea_lemma_2015(pig):
    if pig > 0.5:
        return wc(["green/purple", "purple/green", "green/purplish"], [45, 35, 20])
    return wc(
        [
            "green/purple",
            "green/purplish",
            "green/pale green",
            "green/purple/pale green",
        ],
        [35, 30, 20, 15],
    )


# ══════════════════════════════════════════════════════════════════
# 5. AGRONOMIC MEASUREMENTS
# ══════════════════════════════════════════════════════════════════


def gen_plant_height():
    return nd(187.9, 26.2, 110, 254.5)


def gen_ear_height(ph):
    ratio = np.random.normal(0.568, 0.12)
    ratio = np.clip(ratio, 0.25, 0.80)
    eh = ph * ratio
    return round(float(np.clip(eh, 16, ph - 15)), 1)


def gen_anthesis():
    return int(np.clip(round(np.random.normal(57.0, 4.2)), 47, 66))


def gen_silking(anth):
    asi = int(np.clip(round(np.random.exponential(1.5)), 0, 8))
    return anth + asi


def gen_ear_length():
    return nd(10.9, 1.6, 3.7, 15.5, 1)


def gen_ear_diameter():
    return nd(3.61, 0.37, 2.6, 4.9, 2)


def gen_ear_rows():
    return nd(12.7, 3.63, 9, 40, 1)


def gen_blade_width():
    return nd(8.89, 2.5, 6.4, 15.0, 1)  # most realistic range


def gen_tassel_branches():
    return nd(17.2, 3.57, 4, 40.6, 1)


def gen_tassel_main():
    return nd(35.0, 2.75, 25.1, 42.1, 1)


def gen_tassel_lateral():
    return nd(22.9, 1.96, 11, 29.2, 1)


def gen_row_number_dup():
    return nd(54.9, 3.94, 45, 64, 1)  # col 95 (duplicate, ~DAP-equivalent measurement)


# ══════════════════════════════════════════════════════════════════
# 6. CHEMICAL PROFILE (±-format strings)
# ══════════════════════════════════════════════════════════════════


def pm(val, err):
    return f"{val:.2f} ± {err:.3f}"


def gen_ash():
    return pm(nd(1.53, 0.12, 1.2, 1.9, 4), nd(0.05, 0.04, 0.001, 0.4, 4))


def gen_crude_fat():
    return pm(nd(4.9, 0.8, 3.0, 7.5, 2), nd(0.2, 0.15, 0.01, 0.7, 2))


def gen_crude_fiber():
    return pm(nd(1.8, 0.35, 0.9, 3.0, 2), nd(0.3, 0.25, 0.01, 1.0, 2))


def gen_crude_protein():
    return round(nd(9.5, 1.8, 5.5, 14.0, 2), 2)


def gen_total_carb():
    return pm(nd(76, 6, 65, 90, 2), nd(1.5, 1.0, 0.1, 4.0, 2))


def gen_lysine():
    return pm(nd(0.33, 0.07, 0.15, 0.55, 2), nd(0.05, 0.05, 0.001, 0.20, 3))


def gen_tryptophan():
    return pm(nd(0.048, 0.015, 0.020, 0.090, 3), nd(0.009, 0.007, 0.001, 0.030, 4))


def gen_iron():
    return pm(nd(3.2, 1.5, 1.0, 9.0, 2), nd(0.05, 0.06, 0.001, 0.25, 3))


def gen_zinc():
    return pm(nd(3.4, 0.3, 2.5, 4.2, 2), nd(0.015, 0.015, 0.001, 0.08, 3))


def gen_tpc():
    return pm(nd(5.0, 1.2, 2.5, 8.5, 2), nd(0.35, 0.3, 0.05, 1.5, 2))


def gen_tfc():
    return pm(nd(5.9, 0.8, 4.0, 8.0, 2), nd(0.4, 0.35, 0.05, 1.5, 2))


def gen_tcc():
    return pm(nd(0.55, 0.18, 0.20, 1.0, 2), nd(0.20, 0.18, 0.01, 0.60, 2))


def gen_dpph():
    return pm(nd(45, 15, 20, 75, 2), nd(3.5, 3.0, 0.5, 10.0, 2))


def gen_bcarotene():
    v = nd(0.025, 0.015, 0.001, 0.060, 4)
    e = nd(0.002, 0.002, 0.0001, 0.008, 4)
    return f"{v:.4f}±{e:.4f}"


def gen_starch_amylose(kernel_type):
    """Generate Starch, Amylose, Amylopectin so that Amylose + Amylopectin ≈ Starch."""
    nt = _norm_kt(kernel_type)
    # Waxy (glutinous) → very low amylose
    if "glutin" in nt:
        amylose = nd(3.0, 1.5, 2.0, 8.0, 2)
    else:
        amylose = nd(9.23, 6.46, 2.0, 25.0, 2)
    starch = nd(74.6, 8.38, 53, 92, 2)
    amylopectin = round(float(np.clip(starch - amylose, 37, 90)), 2)
    ratio = round(amylose / max(amylopectin, 0.01), 4)
    return starch, amylose, amylopectin, ratio


# ══════════════════════════════════════════════════════════════════
# 7. DISEASE / STRESS RESISTANCE
# ══════════════════════════════════════════════════════════════════

# Fraction of total records that are *screened* (not null, not "Not yet screened"):
# Waterlogging: (244+198+154+106+5+3+2)/3799 ≈ 18.7%
# Acidic Soil:  542/3799 ≈ 14.3%
# Calcareous:   542/3799 ≈ 14.3%
# Drought:      619/3799 ≈ 16.3%
# Downy Mildew: 380/3799 ≈ 10.0%
# Bacterial:    233/3799 ≈ 6.1%
# Fusarium:     466/3799 ≈ 12.3%
# Corn Borer LDA: screened: (605+138+44+15+10)/3799 ≈ 21.4%
# Corn Weevil:  (446+367+50+22+8+1)/3799 ≈ 23.5%
# Corn Plant Hopper: (252+71+7+4)/3799 ≈ 8.8%

# Among disease cols, 89% of rows have ANY value (incl "Not yet screened");
# so null rate ≈ 11% for those cols.


def _latent_resistance():
    """Draw a latent tolerance score [0,1], skewed toward susceptible."""
    return np.random.beta(1.5, 4.5)


def gen_waterlogging(latent):
    return wc(
        ["Tolerant", "Moderately Tolerant", "Moderately Susceptible", "Susceptible"],
        [
            max(0.01, 0.343 + latent * 0.3),
            max(0.01, 0.216 + latent * 0.1),
            max(0.01, 0.149 - latent * 0.1),
            max(0.01, 0.278 - latent * 0.3),
        ],
    )


def gen_waterlogging_new(wl_rating):
    """Waterlogging New  column (MS/S) — present only when waterlogging is screened."""
    if "Susceptible" in wl_rating and "Moderately" not in wl_rating:
        return "S"
    return "MS"


def gen_adv_roots():
    v = np.random.normal(2.1, 1.32)
    return round(float(np.clip(v, 0, 8)), 1)


def gen_asi_waterlog():
    v = np.random.normal(4.48, 3.63)
    return round(float(np.clip(v, 0, 16)), 1)


def gen_plants_adv_roots():
    return round(float(np.clip(np.random.normal(1.87, 1.23), 0, 9)), 1)


def gen_yield_waterlog():
    v = np.random.normal(0.017, 0.022)
    return round(float(np.clip(v, 0, 0.131)), 4)


def gen_acidic(latent):
    return wc(
        ["Resistant", "Moderately Resistant", "Moderately Susceptible", "Susceptible"],
        [
            max(0.01, 0.054 + latent * 0.2),
            max(0.01, 0.328 + latent * 0.15),
            max(0.01, 0.467 - latent * 0.15),
            max(0.01, 0.151 - latent * 0.2),
        ],
    )


def gen_calcareous(latent):
    return wc(
        ["Resistant", "Moderately Resistant", "Moderately Susceptible", "Susceptible"],
        [
            max(0.01, 0.055 + latent * 0.2),
            max(0.01, 0.342 + latent * 0.15),
            max(0.01, 0.494 - latent * 0.15),
            max(0.01, 0.111 - latent * 0.2),
        ],
    )


def gen_drought(latent):
    return wc(
        ["Resistant", "Moderately Resistant", "Moderately Susceptible", "Susceptible"],
        [
            max(0.01, 0.048 + latent * 0.2),
            max(0.01, 0.308 + latent * 0.15),
            max(0.01, 0.391 - latent * 0.1),
            max(0.01, 0.245 - latent * 0.25),
        ],
    )


def gen_dm(latent):
    return wc(
        ["Resistant", "Moderately Resistant", "Moderately Susceptible", "Susceptible"],
        [
            max(0.01, 0.005 + latent * 0.15),
            max(0.01, 0.055 + latent * 0.15),
            max(0.01, 0.237 - latent * 0.05),
            max(0.01, 0.703 - latent * 0.25),
        ],
    )


def gen_dm_value(dm_rating):
    """Downy Mildew Value is a fraction (0–1) matching the rating."""
    if dm_rating in ("Noy yet Screened", "Not yet screened"):
        return dm_rating
    mapping = {
        "Susceptible": (0.6, 1.0),
        "Moderately Susceptible": (0.3, 0.65),
        "Moderately Resistant": (0.1, 0.35),
        "Resistant": (0.0, 0.15),
    }
    lo, hi = mapping.get(dm_rating, (0.4, 0.8))
    return round(random.uniform(lo, hi), 4)


def gen_bsr(latent):
    return wc(
        ["Resistant", "Moderately Resistant", "Moderately Susceptible", "Susceptible"],
        [
            max(0.01, 0.009 + latent * 0.15),
            max(0.01, 0.099 + latent * 0.15),
            max(0.01, 0.614 - latent * 0.1),
            max(0.01, 0.279 - latent * 0.2),
        ],
    )


def gen_bsr_value(bsr_rating):
    if bsr_rating == "Not yet screened":
        return "Not yet screened"
    mapping = {
        "Susceptible": (0.6, 1.0),
        "Moderately Susceptible": (0.3, 0.65),
        "Moderately Resistant": (0.1, 0.35),
        "Resistant": (0.0, 0.15),
    }
    lo, hi = mapping.get(bsr_rating, (0.4, 0.8))
    return round(random.uniform(lo, hi), 4)


def gen_fusarium(latent):
    return wc(
        [
            "Highly Resistant",
            "Moderately Resistant",
            "Intermediate",
            "Moderately Susceptible",
            "Highly Susceptible",
        ],
        [
            max(0.01, 0.036 + latent * 0.1),
            max(0.01, 0.049 + latent * 0.1),
            max(0.01, 0.215 + latent * 0.05),
            max(0.01, 0.279 - latent * 0.1),
            max(0.01, 0.421 - latent * 0.15),
        ],
    )


def gen_corn_borer_lda(latent):
    return wc(
        [
            "Highly Susceptible",
            "Susceptible",
            "Moderately Resistant",
            "Resistant",
            "Highly Resistant",
        ],
        [
            max(0.01, 0.605 - latent * 0.3),
            max(0.01, 0.251 - latent * 0.1),
            max(0.01, 0.083 + latent * 0.1),
            max(0.01, 0.036 + latent * 0.1),
            max(0.01, 0.019 + latent * 0.1),
        ],
    )


def gen_corn_borer_remarks(lda_rating):
    if "Resistant" in lda_rating:
        return wc(
            [
                "Stalk Resistance for Confirmation",
                "For Confirmation",
                "Need to be reconfirmed",
            ],
            [60, 25, 15],
        )
    return "Not yet screened"


def gen_corn_weevil(latent):
    return wc(
        [
            "Highly Susceptible",
            "Susceptible",
            "Moderately Susceptible",
            "Moderately Resistant",
            "Resistant",
        ],
        [
            max(0.01, 0.446 - latent * 0.25),
            max(0.01, 0.367 - latent * 0.15),
            max(0.01, 0.056),
            max(0.01, 0.022 + latent * 0.1),
            max(0.01, 0.008 + latent * 0.08),
        ],
    )


def gen_cph(latent):
    return wc(
        ["Susceptible", "Moderately Susceptible", "Highly Susceptible", "Resistant"],
        [
            max(0.01, 0.748 - latent * 0.2),
            max(0.01, 0.210 - latent * 0.05),
            max(0.01, 0.021),
            max(0.01, 0.012 + latent * 0.1),
        ],
    )


# ══════════════════════════════════════════════════════════════════
# 8. Bt / GT / GLYPHOSATE
# ══════════════════════════════════════════════════════════════════


def gen_bt_gt():
    """
    Returns (Bt all in data, GT all in, BtGt Positive/Remarks).
    Distributions match source:
      Bt: 78.4% NYT, 19.0% Negative, 2.0% Positive, etc.
      GT: 69.1% NYT, 26.7% Negative, 4.2% Positive
    """
    bt = wc(
        ["Not yet tested", "Negative", "Positive", "Bt/GT POSITIVE"],
        [78.4, 19.0, 2.0, 0.6],
    )
    gt = wc(["Not yet tested", "Negative", "Positive"], [69.1, 26.7, 4.2])
    # Derive summary
    if bt in ("Not yet tested", "Not yet tested") and gt in (
        "Not yet tested",
        "Not yet tested",
    ):
        summary = wc(["both not yet tested", "NYT to BtGt"], [75, 25])
    elif bt == "Negative" and gt == "Negative":
        summary = "BtGt Negative"
    elif bt == "Negative" and gt in ("Not yet tested",):
        summary = "Bt Negative, Gt NYT"
    elif gt == "Negative" and bt in ("Not yet tested",):
        summary = "GT Negative, Bt NYT"
    elif bt == "Positive":
        summary = wc(
            ["Bt positive, Gt not yet tested", "Yes", "Positive Bt only"], [50, 30, 20]
        )
    elif gt == "Positive":
        summary = wc(["Bt NYT, Gt positive", "Gt only"], [60, 40])
    else:
        summary = wc(
            ["both not yet tested", "BtGt Negative", "GT Negative, Bt NYT"],
            [52, 12, 15],
        )
    return bt, gt, summary


# ══════════════════════════════════════════════════════════════════
# 9. PASSPORT / METADATA HELPERS
# ══════════════════════════════════════════════════════════════════

FIRST_NAMES = [
    "Rodrigo",
    "Maria",
    "Jose",
    "Analiza",
    "Juan",
    "Elena",
    "Noel",
    "Leonida",
    "Rudy",
    "Suzana",
    "Benjie",
    "Panso",
    "Sherly",
    "Jackelyn",
    "Mary Ann",
    "Roque",
    "Santos",
    "Felicitas",
    "Benigno",
    "Sinon",
    "Ramona",
    "Efren",
    "Crisanta",
    "Domingo",
    "Gloria",
    "Isidro",
    "Lorna",
    "Macario",
    "Nenita",
    "Oscar",
    "Paz",
    "Quirino",
    "Rosa",
    "Serafin",
    "Teresita",
    "Urbano",
    "Victorina",
    "Wilfredo",
    "Ximena",
    "Yolanda",
    "Zenaida",
    "Adelaida",
    "Bartolome",
    "Caridad",
    "Demetrio",
    "Esperanza",
    "Francisco",
    "Generosa",
    "Herminia",
    "Ignacio",
    "Juanita",
    "Katrina",
    "Lorenzo",
    "Milagros",
    "Nelson",
    "Ofelia",
    "Pedro",
    "Remedios",
    "Salvador",
    "Tomas",
]
LAST_NAMES = [
    "Arsaga",
    "Martinez",
    "Medianero",
    "Labañego",
    "Sandoy",
    "Ritero",
    "Bartolome",
    "Valsado",
    "Capada",
    "Maximo",
    "Kitay",
    "Bacalso",
    "Santos",
    "Reyes",
    "Cruz",
    "Garcia",
    "Lopez",
    "Gonzales",
    "Hernandez",
    "Perez",
    "Ramos",
    "Villanueva",
    "Torres",
    "Flores",
    "Dela Cruz",
    "Mendoza",
    "Rivera",
    "Castillo",
    "Morales",
    "Domingo",
    "Aquino",
    "Pascual",
    "Lim",
    "Tan",
    "Chua",
    "Go",
    "Sy",
    "Uy",
    "Co",
    "Tiu",
    "Calutan",
    "Pangalinawan",
    "Nimeta",
    "Gumila",
    "Beltran",
    "Manato",
    "Merdata",
    "Yabes",
    "Piamonte",
]


def random_name():
    return f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"


def random_phone():
    return f"09{random.randint(100000000, 999999999)}"


COLLECTORS = [
    "GBL, CZC",
    "Rudy Pangalinawan",
    "Melboy  N. Calutan",
    "R.M RAMOS",
    "LEO BELTRAN",
    "Nova Joy Manato & Alden Jovandillo",
    "ELM, LBM, MIJE",
    "RASIB MERDATA/NOHMODIN KADIR",
    "JEAN A. NIMETA",
    "JANE MARDE GUMILA",
    "MAE BELTRAN",
    "MGL, KDM",
    "SRP, GBL",
    "GBL",
    "MGL",
]

INSTITUTES = [
    "-",
    "DA-RFO IV-A",
    "BIARC",
    "DA-RFO I",
    "DA-BIAR",
    "DA-Research Division",
    "DA-RFO IV-B",
    "ARMMARC",
    "DA-RFU XII",
    "DMLRRS",
    "TRENTO RES",
    "IPB-UPLB",
    "DA-CVARP",
    "DA-RFO VI",
    "DA-RFO VII",
    "DA-RFU IX",
]

DATE_RECEIVED = [
    "April 29, 2019",
    "October 5, 2018",
    "May 31, 2018",
    "January 7, 2019",
    "September 4, 2018",
    "August 16, 2018",
    "May 26, 2018",
    "11/15/2017",
    "December 6, 2018",
    "March 20, 2019",
    "June 15, 2018",
    "February 8, 2019",
    "July 12, 2019",
    "August 30, 2019",
    "October 19, 2020",
    "Dec. 6, 2020",
    "March 4, 2021",
    "June 10, 2021",
    "November 5, 2021",
    "February 2022",
    "May 2022",
    "October 2022",
    "January 2023",
    "February 2023",
]

DATES_COLLECTION = [
    "-",
    "March 07, 2018",
    "August 25, 2015",
    "APRIL 11, 2018",
    "July 2018",
    "August 2018",
    "September 2018",
    "November 2018",
    "January 2019",
    "March 2019",
    "June 2019",
    "September 2019",
    "2017-08-03",
    "2016-09-17",
    "2015-08-25",
    "January 2020",
    "June 2021",
    "October 2021",
    "2022-06-29",
    "2022-10-15",
]

PLANTING_DATES = [
    "-",
    "May",
    "June",
    "July",
    "March",
    "April",
    "January",
    "November",
    "MAY 2016",
    "MAY 2015",
    "NOVEMBER 2015",
    "FEBRUARY 2016",
    "JANUARY 2016",
    "MARCH 2017",
    "MAY 2018",
    "JUNE 2018",
    "2018-04-01 00:00:00",
    "2018-03-01 00:00:00",
]

HARVEST_DATES = [
    "-",
    "August",
    "September",
    "October",
    "February",
    "March",
    "April",
    "July",
    "AUGUST 2016",
    "FEBRUARY 2016",
    "AUGUST 2015",
    "SEPTEMBER 2015",
    "AUGUST 2017",
    "SEPTEMBER 2018",
    "2018-07-01 00:00:00",
    "OCTOBER 2018",
    "NOVEMBER 2018",
]

COLLECTING_SOURCES_W = [
    ("FARMLAND", 736),
    ("farmland", 219),
    ("Farmland", 93),
    ("Backyard", 40),
    ("BACKYARD", 38),
    ("backyard/home garden", 25),
    ("house", 18),
    ("FARM STORE", 14),
    ("farm store/threshing place", 9),
]

GENETIC_STATUS_W = [
    ("NATIVE", 397),
    ("FARMLAND", 216),
    ("NATIVE/LANDRACE CULTIVAR", 202),
    ("Native/ landrace cultivars", 156),
    ("native/landrace cultivar", 100),
    ("native/landrace variety", 72),
    ("IMPROVED OPV", 22),
    ("native", 22),
]

TOPOGRAPHY_W = [
    ("UPLAND PLAIN", 174),
    ("upland plain", 129),
    ("LOWLAND PLAIN", 141),
    ("lowland plain", 59),
    ("HILLY", 133),
    ("hilly", 95),
    ("LEVEL PLAIN", 115),
    ("level plain", 52),
    ("FLOOD PLAIN", 92),
    ("MOUNTAINOUS", 66),
    ("mountainous", 18),
]

SITE_W = [
    ("-", 958),
    ("LEVEL", 215),
    ("NON-RANDOM", 180),
    ("SLOPE", 145),
    ("level", 96),
    ("slope", 67),
    ("RANDOM", 33),
    ("UNDULATING", 20),
    ("PLAIN", 17),
]

SOIL_TEXTURE_W = [
    ("SANDY LOAM", 301),
    ("CLAY LOAM", 180),
    ("CLAYEY", 125),
    ("clayey", 109),
    ("SANDY", 95),
    ("sandy", 95),
    ("sandy loam", 74),
    ("clay loam", 48),
    ("SILTY", 44),
    ("LOAMY", 30),
    ("loamy", 20),
]

SOIL_COLOR_W = [
    ("LIGHT BROWN", 307),
    ("DARK BROWN", 251),
    ("light brown", 135),
    ("BLACK", 133),
    ("dark brown", 93),
    ("black", 56),
    ("GREY", 32),
    ("grey", 13),
    ("RED", 14),
]

STONINESS_W = [
    ("LOW", 268),
    ("MODERATE", 203),
    ("ALMOST NONE", 164),
    ("almost none", 131),
    ("moderate", 109),
    ("low", 66),
    ("MEDIUM", 32),
    ("medium", 20),
    ("HIGH", 18),
]

FARMING_W = [
    ("CROP ROTATION", 219),
    ("crop rotation", 139),
    ("INTERCROPPING/MULTIPLE CROPPING", 168),
    ("intercropping/ multiplecropping", 37),
    ("SHIFTING/KAINGIN", 34),
    ("HILL/SLOPE TERRACING", 43),
    ("monocropping", 21),
    ("crop rotaion", 26),
]

SOWING_W = [
    ("FURROW SEEDING", 239),
    ("furrow seeding", 145),
    ("DIRECT SEEDING/DIBBLING", 144),
    ("direct-seeding/dibbling", 85),
    ("direct seeding/dibbling", 47),
    ("DIRECT-SEEDING/DIBBLING", 79),
    ("furrowing", 28),
]

IRRIGATION_W = [
    ("rainfed", 379),
    ("RAINFED", 328),
    ("Rainfed", 10),
    ("SHALLOW/DEEP WELL/PUMP", 88),
    ("SPRING/RIVER", 44),
    ("spring/ river", 10),
    ("not indicated", 8),
]

DRAINAGE_W = [
    ("GOOD", 397),
    ("good", 202),
    ("MODERATE", 148),
    ("POOR", 115),
    ("moderate", 96),
    ("poor", 34),
    ("not indicated", 22),
    ("EXCESSIVE", 8),
]

CROP_SEQ_W = [
    ("corn-corn", 60),
    ("CORN-CORN", 57),
    ("RICE-CORN", 56),
    ("CORN", 45),
    ("CORN-VEGETABLES", 31),
    ("CORN-RICE", 14),
    ("corn-rice", 11),
    ("corn-Peanut", 9),
    ("CORN-PEANUT", 8),
    ("rice-corn", 9),
]

FERTILIZER_W = [
    ("21-0-0", 27),
    ("N/A", 25),
    ("INORGANIC", 19),
    ("No fertilizer applied", 17),
    ("NO FERTILIZER", 16),
    ("16-20-0", 13),
    ("urea", 11),
    ("inorganic", 10),
]

PEST_W = [
    ("MANUAL WEEDING", 51),
    ("HAND WEEDING", 31),
    ("N/A", 26),
    ("hand weeding", 26),
    ("No pesticide applied", 16),
    ("gunahan", 15),
    ("tudling", 11),
    ("NO CHEMICAL USED", 11),
    ("SPOT WEEDING", 10),
]

USAGE_W = [
    ("FOOD", 57),
    ("consumption", 51),
    ("food", 42),
    ("Food", 39),
    ("FOR FOOD", 27),
    ("animal feed, consumption", 10),
    ("sell", 10),
    ("consumption and commercial", 10),
    ("FOR FOOD/MARKET", 8),
]

SAMPLE_TYPE_W = [
    ("WHOLE EARS/COBS", 229),
    ("SHELLED GRAINS/KERNEL", 142),
    ("WHOLE EARS", 95),
    ("wole ears/cob", 91),
    ("SHELLED GRAINS/KERNELS", 76),
    ("WHOLE EARS/COB", 70),
    ("whole ears", 66),
    ("SHELLED GRAINS", 58),
]

GENETIC_STATUS_COLLECT_MAP = {
    "FARMLAND": "FARMLAND",
    "farmland": "native/landrace cultivar",
    "Farmland": "Native/ landrace cultivars",
    "Backyard": "NATIVE",
    "BACKYARD": "NATIVE/LANDRACE CULTIVAR",
}

CEREALS_BREEDING_W = [
    ("#N/A", 458),
    ("Seed Regeneration (Half-Sib), Pili drive, Planted 2018 Q4, 4 rows", 316),
    ("Seed Regeneration(Half-sib), D3 East, planted 2018 Q3, 5 rows", 229),
    (
        "Seed Regeneration (Full-sib) and Characterization, Pili drive A10, Planted 2019 Q3, 4 rows",
        17,
    ),
    ("Seed Regeneration (Full-sib), Pili drive A10, Planted 2019 Q3, 4 rows", 3),
]

LOCAL_NAMES_W = [
    ("Tiniguib", 544),
    ("Lagkitan", 161),
    ("Pilit", 138),
    ("Mimis", 102),
    ("Native", 99),
    ("Malagkit", 84),
    ("Takuro", 67),
    ("Pulutan", 66),
    ("Native Corn", 60),
    ("Kalimpos", 60),
    ("No Name Indicated", 45),
    ("Calimpus", 39),
    ("Diket", 38),
    ("Tiniguib B", 32),
    ("Pula", 28),
    ("Mais Pula", 25),
    ("Putian", 22),
    ("Seniorita", 18),
    ("Lawaan", 15),
    ("Batik", 12),
    ("Toledo", 8),
    ("Basay", 7),
    ("Butil Mais", 5),
    ("Mais Dilaw", 5),
    ("Tiniguib A", 5),
    ("Tiniguib C", 4),
    ("Tiniguib D", 3),
    ("Kalamay", 10),
    ("Pinalit", 8),
    ("Panggas", 6),
    ("Mais Puti", 15),
    ("Lusay", 9),
    ("Puso", 7),
    ("Takuro Puti", 6),
    ("Tikoy", 5),
    ("Dinugan", 4),
    ("Malagkiting Mais", 10),
]

MEANING_OF_NAME_W = [
    ("-", 1410),
    ("MALAGKIT", 38),
    ("Native", 11),
    ("Karaan", 8),
    ("DIKET", 8),
    ("GLUTINOUS", 7),
    ("LAGKITAN", 7),
    ("HONEY CORN", 3),
    ("LATE MATURING", 2),
    ("STICKY CORN", 3),
    ("OLD VARIETY", 2),
    ("WHITE CORN", 4),
    ("RED CORN", 2),
]

SAMPLING_METHOD_W = [
    ("RANDOM", 613),
    ("random", 235),
    ("SHELLED", 110),
    ("non random", 94),
    ("NON-RANDOM", 93),
    ("non-random", 81),
    ("WHOLE EARS", 72),
]

EAR_SHAPE_W = [
    ("Tapered", 174),
    ("Conical", 104),
    ("Cylindrical", 36),
    ("No data", 47),
    ("Regular", 1),
]
GRAIN_ROW_ARR_W = [("Regular", 306), ("No data", 50), ("Irregular", 4), ("Spiral", 2)]
SHAPE_OF_GRAIN_W = [("Round", 255), ("Elliptical", 57), ("No data", 49)]

NPGRL_NOTES_FORMAT = "CGUARD{year}-{num:05d}"

KEY_DESC_W = [
    ("-", 1320),
    ("FARMLAND COLLECTION OF SAMPLE", 10),
    ("FARMLAND SAMPLES", 7),
    ("good eating quality", 7),
    ("Good milling recovery", 5),
    ("FARM SAMPLE COLLECTION", 5),
    ("SOFT", 4),
    ("GLUTINOUS", 4),
    ("BIG COBS AND KERNELS", 4),
    ("EARLY MATURING", 3),
    ("LATE MATURING", 2),
    ("THICK HUSKS", 2),
    ("GOOD STALK STRENGTH", 2),
]

DIALECT_FALLBACK = [
    "-",
    "CEBUANO",
    "BISAYA",
    "TAGALOG",
    "VISAYAN",
    "BICOL",
    "ILOCO",
    "IBANAG, ILOCANO",
    "MAGUINDANAON",
    "SUBANEN",
    "WARAYWARAY",
]


# ══════════════════════════════════════════════════════════════════
# 10. MAIN RECORD GENERATOR
# ══════════════════════════════════════════════════════════════════


def _ww(dist):
    """Helper: sample from (value, weight) list."""
    vals, ws_ = zip(*dist)
    return wc(list(vals), list(ws_))


def generate_record(apn: int) -> dict:
    # ── Geography ──
    regions, rw = zip(*[(r, d["w"]) for r, d in REGION_DATA.items()])
    region = wc(list(regions), list(rw))
    rd = REGION_DATA[region]
    prov_pool = rd["provinces"]
    province = wc([p for p, _ in prov_pool], [w for _, w in prov_pool])
    munis = PROVINCE_MUNICIPALITIES.get(
        province, ["Poblacion", "San Jose", "San Francisco"]
    )
    town = random.choice(munis)
    barangay = random.choice(BARANGAYS)
    sitio = wc(SITIOS, [3 if s == "-" else 1 for s in SITIOS])
    dialect = (
        random.choice(rd["dialects"])
        if random.random() < 0.80
        else random.choice(DIALECT_FALLBACK)
    )

    # ── Kernel genetics ──
    kernel_type = _ww(KERNEL_TYPES_W)
    kernel_color = gen_kernel_color(kernel_type)
    pigment = get_pigment(kernel_color)
    aleurone = gen_aleurone(kernel_color)
    endosperm = gen_endosperm(kernel_color, kernel_type)
    kernel_color_std = gen_kernel_color_std(kernel_color)
    cob_color = gen_cob_color(kernel_color)
    grain_type = gen_grain_type(kernel_type)

    # ── Local name / designations ──
    local_name = _ww(LOCAL_NAMES_W)
    meaning = _ww(MEANING_OF_NAME_W)
    raphy_des = (
        f"{local_name.replace(' ', '')}{barangay.replace(' ', '').replace('.', '')}"
    )

    # ── Latent resistance ──
    latent = _latent_resistance()

    # ── Screening flags (control which records have measured data) ──
    has_passport = random.random() < 0.424
    has_agronomic = random.random() < 0.095  # cols 80-97 fill rate
    has_2015_morph = random.random() < 0.015  # cols 66-79 fill rate
    has_chem = random.random() < 0.038  # chemical profile cols
    has_wl_trial = random.random() < 0.038  # waterlogging trial cols 120-124
    has_disease = random.random() < 0.89  # disease screening cols have any value
    has_wl_screen = has_disease and (random.random() < 0.187 / 0.89)
    has_acid = has_disease and (random.random() < 0.143 / 0.89)
    has_calc = has_disease and (random.random() < 0.143 / 0.89)
    has_drought = random.random() < 0.163
    has_dm = has_disease and (random.random() < 0.100 / 0.89)
    has_bsr = has_disease and (random.random() < 0.061 / 0.89)
    has_fer = has_disease and (random.random() < 0.123 / 0.89)
    has_cb_lda = has_disease and (random.random() < 0.214 / 0.89)
    has_cw = has_disease and (random.random() < 0.235 / 0.89)
    has_cph = has_disease and (random.random() < 0.088 / 0.89)

    # ── Agronomic measurements ──
    plant_height = gen_plant_height() if has_agronomic else None
    ear_height = gen_ear_height(plant_height) if has_agronomic else None
    anthesis_dap = gen_anthesis() if has_agronomic else None
    silking_dap = gen_silking(anthesis_dap) if has_agronomic else None
    ear_len = gen_ear_length() if has_agronomic else None
    ear_diam = gen_ear_diameter() if has_agronomic else None
    ear_rows = gen_ear_rows() if has_agronomic else None
    blade_w = gen_blade_width() if has_agronomic else None
    tassel_br = gen_tassel_branches() if has_agronomic else None
    tassel_main = gen_tassel_main() if has_agronomic else None
    tassel_lat = gen_tassel_lateral() if has_agronomic else None
    rows_dup = gen_row_number_dup() if has_agronomic else None

    # Ear shape (col 80) and derived cols only when agronomic
    if has_agronomic:
        ear_shape = _ww([(v, w) for v, w in EAR_SHAPE_W if v != "No data"])
        grain_row = wc(
            ["Regular", "Irregular", "Spiral"],
            [0.60, 0.35, 0.05] if ear_shape == "Conical" else [0.90, 0.07, 0.03],
        )
        shape_grain = _ww([(v, w) for v, w in SHAPE_OF_GRAIN_W if v != "No data"])
    else:
        ear_shape = grain_row = shape_grain = None

    # ── 2015 morphological ──
    ND = "NO DATA"
    stem_c_2015 = gen_stem_color_2015(pigment) if has_2015_morph else ND
    sheath_2015 = gen_sheath_2015(pigment) if has_2015_morph else ND
    spikelet_2015 = gen_spikelet_2015(pigment) if has_2015_morph else ND
    palea_2015 = gen_palea_lemma_2015(pigment) if has_2015_morph else ND
    anther_2015 = gen_anther_2015(pigment) if has_2015_morph else ND
    silk_2015 = gen_silk_color_2015(pigment) if has_2015_morph else ND
    leaves_fn = round(nd(13.2, 0.9, 11, 15, 2), 2) if has_2015_morph else ND
    leaf_angle_v = round(nd(78.0, 3.5, 70, 90, 2), 2) if has_2015_morph else ND
    leaves_ear = round(nd(5.5, 0.6, 4, 7, 1), 1) if has_2015_morph else ND
    crop_stand = int(nd(13.5, 3.5, 5, 16, 0)) if has_2015_morph else ND
    spad_val = round(nd(46, 4.5, 35, 58, 2), 2) if has_2015_morph else ND
    anth_2015_v = int(nd(48.2, 2.3, 38, 53)) if has_2015_morph else ND
    silk_2015_v = int(nd(48.5, 2.0, 40, 55)) if has_2015_morph else ND
    asi_2015_v = int(nd(50.4, 3.2, 42, 58)) if has_2015_morph else ND  # col 77

    # ── Chemical profile ──
    if has_chem:
        ash_v = gen_ash()
        fat_v = gen_crude_fat()
        fiber_v = gen_crude_fiber()
        prot_v = gen_crude_protein()
        carb_v = gen_total_carb()
        lys_v = gen_lysine()
        tryp_v = gen_tryptophan()
        starch_v, amyl_v, amylop_v, ratio_v = gen_starch_amylose(kernel_type)
        iron_v = gen_iron()
        zinc_v = gen_zinc()
        dpph_v = gen_dpph()
        tpc_v = gen_tpc()
        tfc_v = gen_tfc()
        tcc_v = gen_tcc()
        bcaro_v = gen_bcarotene()
        fat_rank = random.randint(1, 83)
        dpph_rank = random.randint(1, 101)
        bcaro_rank = random.randint(1, 83)
    else:
        ash_v = fat_v = fiber_v = prot_v = carb_v = lys_v = tryp_v = None
        starch_v = amyl_v = amylop_v = ratio_v = iron_v = zinc_v = None
        dpph_v = tpc_v = tfc_v = tcc_v = bcaro_v = None
        fat_rank = dpph_rank = bcaro_rank = None

    # ── Waterlogging trial ──
    if has_wl_trial:
        wl_new = wc(["MS", "S"], [73.6, 26.4])
        adv_root = gen_adv_roots()
        asi_wl = gen_asi_waterlog()
        pl_adv = gen_plants_adv_roots()
        yield_wl = gen_yield_waterlog()
    else:
        wl_new = adv_root = asi_wl = pl_adv = yield_wl = None

    # ── Disease screening ──
    wl_rating = (
        gen_waterlogging(latent)
        if has_wl_screen
        else ("Not yet screened" if has_disease else None)
    )
    acid_rat = (
        gen_acidic(latent)
        if has_acid
        else ("Not yet screened" if has_disease else None)
    )
    calc_rat = (
        gen_calcareous(latent)
        if has_calc
        else ("Not yet screened" if has_disease else None)
    )
    drought_r = gen_drought(latent) if has_drought else None

    dm_rating = (
        gen_dm(latent) if has_dm else ("Noy yet Screened" if has_disease else None)
    )
    dm_val = (
        gen_dm_value(dm_rating)
        if has_dm
        else (
            wc(["Noy yet Screened", "Not yet screened"], [89, 11])
            if has_disease
            else None
        )
    )
    bsr_rating = (
        gen_bsr(latent) if has_bsr else ("Not yet screened" if has_disease else None)
    )
    bsr_val = (
        gen_bsr_value(bsr_rating)
        if has_bsr
        else ("Not yet screened" if has_disease else None)
    )

    if has_fer:
        fer_rating = gen_fusarium(latent)
    else:
        fer_rating = "Not yet screened" if random.random() < 0.86 else None

    if has_cb_lda:
        cb_lda = gen_corn_borer_lda(latent)
        cb_sda_r = wc(["Not yet screened", "Resistant", "Susceptible"], [97, 2, 1])
        cb_rem = gen_corn_borer_remarks(cb_lda)
    else:
        cb_lda = (
            wc(["#N/A", "0", "Highly Susceptible"], [60, 22, 18])
            if has_disease
            else None
        )
        cb_sda_r = "Not yet screened" if has_disease else None
        cb_rem = "Not yet screened" if has_disease else None

    cw_rating = (
        gen_corn_weevil(latent)
        if has_cw
        else ("Not yet screened" if has_disease else None)
    )
    cph_rating = (
        gen_cph(latent) if has_cph else ("Not yet screened" if has_disease else None)
    )

    # ── Bt / GT ──
    bt_val, gt_val, btgt_summary = gen_bt_gt()

    # ── Region number format ──
    region_code = region.replace(" ", "").upper()
    if random.random() < 0.05:
        reg_num = wc(
            ["No Reg #, old collection", "NO REG #, OLD COL", "no region #"],
            [67, 25, 8],
        )
    else:
        reg_num = f"CGUARD-{region_code}-{random.randint(1, 400):04d}"

    # ── CGUARD NPGRL number ──
    npgrl_note = (
        wc(["-", "CGUARD2016-0" + str(random.randint(100, 999))], [88, 12])
        if has_passport
        else None
    )

    # ── Cereals breeding ──
    cereals_note = _ww(CEREALS_BREEDING_W) if random.random() < 1023 / 3799 else None

    # ── Misc passport fields ──
    coll_src = _ww(COLLECTING_SOURCES_W) if has_passport else None
    gen_status = (
        GENETIC_STATUS_COLLECT_MAP.get(coll_src, _ww(GENETIC_STATUS_W))
        if coll_src
        else None
    )
    topog = _ww(TOPOGRAPHY_W) if has_passport else None
    site_v = _ww(SITE_W) if has_passport else None
    drainage_v = _ww(DRAINAGE_W) if has_passport else None
    irrigation = _ww(IRRIGATION_W) if has_passport else None

    # Soil texture/color correlate with topography
    if topog and "FLOOD" in topog.upper():
        soil_t = wc(["CLAY LOAM", "CLAYEY", "clayey"], [50, 30, 20])
    elif topog and "UPLAND" in topog.upper():
        soil_t = wc(["SANDY LOAM", "SANDY", "sandy loam", "LOAMY"], [40, 25, 20, 15])
    else:
        soil_t = _ww(SOIL_TEXTURE_W) if has_passport else None

    soil_c = _ww(SOIL_COLOR_W) if has_passport else None
    stonin = _ww(STONINESS_W) if has_passport else None
    farming = _ww(FARMING_W) if has_passport else None
    sowing = _ww(SOWING_W) if has_passport else None
    crop_s = _ww(CROP_SEQ_W) if has_passport else None
    fertil = _ww(FERTILIZER_W) if has_passport else None
    pest = _ww(PEST_W) if has_passport else None
    usage = _ww(USAGE_W) if has_passport else None
    samp_t = _ww(SAMPLE_TYPE_W) if has_passport else None
    samp_m = _ww(SAMPLING_METHOD_W) if has_passport else None
    lat_v = (
        (f"{random.uniform(6.5, 18.5):.8f}" if random.random() < 0.44 else "-")
        if has_passport
        else None
    )
    lon_v = (
        (f"{random.uniform(118.0, 126.5):.8f}" if random.random() < 0.44 else "-")
        if has_passport
        else None
    )
    alt_v = (
        (
            wc(
                [
                    "-",
                    f"{random.randint(10, 600)} MASL",
                    f"{random.randint(10, 600)} M",
                ],
                [74, 15, 11],
            )
        )
        if has_passport
        else None
    )
    yrs_pos = (
        (
            wc(
                ["-", "5", "10", "15", "20", "25", "30", "40", "50", "2", "3", "8"],
                [71, 4, 5, 2, 5, 1, 2, 2, 3, 3, 1, 2],
            )
        )
        if has_passport
        else None
    )
    yrs_plt = (
        (
            wc(
                ["-", "2", "3", "5", "10", "20", "30", "40", "4", "7", "15"],
                [84, 3, 2, 3, 4, 2, 1, 1, 2, 1, 1],
            )
        )
        if has_passport
        else None
    )
    key_desc = _ww(KEY_DESC_W) if has_passport else None
    collector = (
        wc(COLLECTORS + ["-"], [*[15] * len(COLLECTORS), 10]) if has_passport else None
    )
    institute = (
        wc(
            INSTITUTES,
            [115, 67, 29, 23, 18, 17, 17, 15, 14, 14, 12, 12]
            + [5] * (len(INSTITUTES) - 12),
        )
        if has_passport
        else None
    )
    grower = (
        random_name()
        if has_passport and random.random() < 0.70
        else ("-" if has_passport else None)
    )
    contact = (
        (random_phone() if random.random() < 0.15 else "-") if has_passport else None
    )
    date_rec = random.choice(DATE_RECEIVED) if random.random() < 0.286 else None
    date_col = random.choice(DATES_COLLECTION) if has_passport else None
    date_pl = random.choice(PLANTING_DATES) if has_passport else None
    date_har = random.choice(HARVEST_DATES) if has_passport else None
    crop_v = (
        wc(
            ["-", "CORN", "Corn", "corn", "Native", "NATIVE WHITE CORN"],
            [94, 3, 1, 1, 1, 1],
        )
        if has_passport
        else None
    )
    dialect_v = dialect if has_passport else None
    passport_v = "Yes" if has_passport else None

    # ── Build the full record in EXACT column order ──
    rec = {
        "APN": apn,
        "Region": region,
        "CGUARD N": f"CGUARD N{apn}",
        "Kernel Color": kernel_color,
        "Kernel Type": kernel_type,
        "Local Name ": local_name,  # trailing space preserved
        "identified resistant/tolerant": maybe(
            67 / 3799 * 3,
            _ww(
                [
                    ("ACB", 10),
                    ("FER", 9),
                    ("BSR", 8),
                    ("CW", 8),
                    ("Drought", 5),
                    ("DM", 4),
                    ("Calcareous Soil", 4),
                    ("CPH", 4),
                    ("Waterlogging", 3),
                ]
            ),
        ),
        "Varietal List 1st batch": maybe(100 / 3799 * 3, "Yes"),
        "Date received by Cereals": date_rec,
        "Region #": reg_num,
        "Duplication": maybe(519 / 3799 * 2, random.randint(1, 131)),
        "_unnamed_": None,
        "Remarks to data searched": maybe(
            199 / 3799 * 3,
            wc(
                ["2 entry", "not in the e passport", "dup Reg #", "no local name"],
                [51, 14, 14, 4],
            ),
        ),
        "Remarks": maybe(
            756 / 3799 * 2,
            wc(
                [
                    "NO E-COPY",
                    "No Passport Data",
                    "w/ passport Data",
                    "DUPLICATE",
                    "Duplicate Reg #",
                    "No Passport Data, to identify reg #",
                    "no Passport",
                ],
                [130, 83, 56, 45, 25, 18, 17],
            ),
        ),
        "Other info": maybe(
            101 / 3799 * 3,
            wc(
                [
                    "to check record book",
                    "Tiniguib Strains",
                    "to search more info for confirmation",
                    "IVA was informed re this duplicates",
                ],
                [50, 20, 10, 5],
            ),
        ),
        "Region GB #1": maybe(
            702 / 3799 * 2,
            wc(
                ["-", "005-10", "002-10", "Diket", "003-10", "003-11"],
                [67, 14, 14, 12, 11, 11],
            ),
        ),
        "Region GB #2": maybe(
            2063 / 3799,
            wc(
                ["-", "120-08 F", "143-09", "137-09", "218-10", "23"],
                [1298, 3, 3, 3, 3, 3],
            ),
        ),
        "Collector": collector,
        "Institute": institute,
        "Date of Collection": date_col,
        "Province": province,
        "Town/Municipality": town,
        "Barangay": barangay,
        "Sitio": sitio,
        "Latitude": lat_v,
        "Longitude": lon_v,
        "Altitude": alt_v,
        "Crop": crop_v,
        "Meaning Of Name": meaning,
        "Passport": passport_v,
        "Dialect": dialect_v,
        "Date Planted": date_pl,
        "Date Harvested": date_har,
        "Source/Grower Name": grower,
        "Contact Information": contact,
        "Collecting Source": coll_src,
        "Sample Type": samp_t,
        "Genetic Status": gen_status,
        "Sampling Method": samp_m,
        "Topography": topog,
        "Site": site_v,
        "Soil Texture": soil_t,
        "Soil Color": soil_c,
        "Stoniness": stonin,
        "Farming/Cultural Practice": farming,
        "Sowing": sowing,
        "Irrigation/Water Sources": irrigation,
        "Drainage": drainage_v,
        "Crop Sequence": crop_s,
        "Fertilizer Mgt Practices": fertil,
        "Pest/Weed Mgt Practices": pest,
        "Usage": usage,
        "Key Description Of The Cultivar": key_desc,
        "Years In Possession": yrs_pos,
        "Number Of Years Planted": yrs_plt,
        "Maturity": wc(
            [
                "-",
                "90 DAP",
                "80-85 DAP",
                "90-95 DAP",
                "75 DAP AT HARVEST",
                "60 DAP",
                "60-70 DAP",
                "3 MONTHS OR 90 DAYS TO MATURE",
            ],
            [995, 1, 1, 1, 1, 1, 1, 1],
        )
        if has_passport
        else None,
        "Other Info (Usage, Years Of Posession, Others…)": maybe(
            94 / 3799 * 3,
            wc(
                [
                    "-",
                    "70 YEARS IN POSSESION",
                    "50 YEARS IN POSSESION",
                    "SELL / 60 YEARS",
                    "CONSUMPTION,SELL/ 30 YEARS",
                ],
                [1501, 3, 3, 2, 2],
            ),
        ),
        "Description Of Photos Attached": maybe(
            6 / 3799 * 3,
            wc(
                [
                    "NEWLY HARVESTED",
                    "PURPLE IN COLOR",
                    "WHITE AND PURPLE IN COLOR",
                    "SMALL COBS, POINTED TIP",
                ],
                [1, 1, 1, 1],
            ),
        ),
        "Passport Data Source": maybe(
            78 / 3799 * 3, wc(["-", "E-COPY", "hardcopy"], [1500, 67, 7])
        ),
        "Other Info": maybe(
            35 / 3799 * 3,
            wc(
                ["-", "GREEN SHEET:DUPLICATE IDS, DIFFERENT INFO", "with white corn"],
                [1530, 18, 3],
            ),
        ),
        "NPGRL .": npgrl_note,
        "CGUARD # (From NPGRL old notes)": maybe(
            211 / 3799 * 2, random.randint(17, 386)
        ),
        "Raphy's Designation": raphy_des if random.random() < 0.81 else None,
        "Photograph": None,
        "ALEURONE COLOR": aleurone,
        "ENDOSPERM COLOR": endosperm,
        "Mean number functional leaves (45 DAS) 2015": leaves_fn,
        "Mean leaf angle (leaf above the ear; 45 DAS) 2015": leaf_angle_v,
        "Stem color of the internode below the ear 2015": stem_c_2015,
        "Leaf sheath color below ear leaf 2015": sheath_2015,
        "Mean number of leaves above the top-most ear 2015": leaves_ear,
        "Spikelet base coloration (greenish or purple) 2015": spikelet_2015,
        "Palea and lemma coloration 2015": palea_2015,
        "Anther color 2015": anther_2015,
        "Silk color 2015": silk_2015,
        "Days to 50% anthesis 2015": anth_2015_v,
        "Days to 50% silking 2015": silk_2015_v,
        "50% Anthesis-Silking 2016": asi_2015_v,
        "Crop stand at 3m-row 2015": crop_stand,
        "SPAD of topmost open leaf (45 DAS) 2015": spad_val,
        "Ear shape": ear_shape,
        "Grain row arrangement": grain_row,
        "Type of grain (in middle third of ear)": grain_type if has_agronomic else None,
        "Shape of grain": shape_grain,
        "Kernel color": kernel_color_std if has_agronomic else None,
        "Cob color": cob_color if has_agronomic else None,
        "Width of blade (leaf of upper ear; cm)": blade_w,
        "Plant Height (cm)": plant_height,
        "Ear Height (cm)": ear_height,
        "Number of primary tassel branches": tassel_br,
        "Length of tassel main axis above lowest main branch (cm)": tassel_main,
        "Length of main axis above highest lateral branch (cm)": tassel_lat,
        "Ear length (without husk; cm)": ear_len,
        "Ear diameter (in middle; cm)": ear_diam,
        "Number of rows of grain": ear_rows,
        "Number of rows of grain_1": rows_dup,
        "Time of anthesis (DAP)": anthesis_dap,
        "Time of silk Emergence (DAP)": silking_dap,
        "Ash": ash_v,
        "Crude fat": fat_v,
        "Crude fat Rank  (Julie)": fat_rank,
        "Crude protein": prot_v,
        "Crude fiber": fiber_v,
        "Total carbohydrates": carb_v,
        "Lysine": lys_v,
        "Tryptophan": tryp_v,
        "Starch": starch_v,
        "Amylose": amyl_v,
        "Amylopectin": amylop_v,
        "Amylose/Amylopectin": ratio_v,
        "Iron": iron_v,
        "Zinc": zinc_v,
        "DPPH": dpph_v,
        "DPPH  rank (Julie)": dpph_rank,
        "TPC": tpc_v,
        "TFC": tfc_v,
        "TCC": tcc_v,
        "AVERAGE B-CAROTENE (mg/g)": bcaro_v,
        "Rank": bcaro_rank,
        "Waterlogging": wl_rating,
        "Waterlogging New ": wl_new,  # trailing space preserved
        "Ave. # of Adv Roots": adv_root,
        "ASI": asi_wl,
        "# of plants w Adv Roots": pl_adv,
        "Yield": yield_wl,
        "Acidic Soil": acid_rat,
        "Calcareous Soil": calc_rat,
        "Drought": drought_r,
        "Downy Mildew Value": dm_val,
        "Downy Mildew": dm_rating,
        "Bacterial Stalk Rot": bsr_rating,
        "Bacterial Stalk Rot Value": bsr_val,
        "Fusarium Ear Rot": fer_rating,
        "Corn Borer (LDA)": cb_lda,
        "Corn Borer (SDA)": cb_sda_r,
        "Corn Borer Remarks": cb_rem,
        "Corn Weevil": cw_rating,
        "Corn Plant Hopper": cph_rating,
        "NPGRL": None,
        "Cereals' Breeding": cereals_note,
        "S1 extraction 2018 Q3": maybe(
            18 / 3799 * 3, "S1 extraction, Trunca D3 East, 2018 Q3, 10 rows"
        ),
        "S1 extraction 2018 Q31": maybe(
            25 / 3799 * 3, "S1 extraction, A8 Pili Drive, 2018 Q4, 10 rows"
        ),
        "C1,S1 extraction, 2018 Q4": maybe(
            2 / 3799 * 3, "C1,S1 extraction, A8 Pili Drive, 2018 Q4, 10 rows"
        ),
        "additional S1 extraction, A8 Pili Drive, 2018 Q4, 10 rows": maybe(
            7 / 3799 * 3, "additional S1 extraction, A8 Pili Drive, 2018 Q4, 10 rows"
        ),
        "S2 Status: C0 S1's Gen:DM": maybe(
            6 / 3799 * 5, f"{random.randint(20, 120)}: submitted for Pop imp"
        ),
        "S2 Status: C0 S1's Gen:BSR": maybe(
            3 / 3799 * 5, f"{random.randint(50, 160)}:Sub"
        ),
        "Remarks1": maybe(
            5 / 3799 * 5,
            wc(
                [
                    "non-native",
                    "S2 lines submitted to lab for selection",
                    "no feedback yet",
                ],
                [2, 1, 1],
            ),
        ),
        "S2 Status: C0 S1's Gen:ACB": maybe(
            11 / 3799 * 5, f"{random.randint(28, 90)} *"
        ),
        "S2 Status: C0 S1's Gen:CW": maybe(
            11 / 3799 * 5,
            wc(
                ["CW", "117 submitted: done recom", "62 submitted; done recombination"],
                [9, 1, 1],
            ),
        ),
        "S2 Status: C0 S1's Gen: Calcareous Soil": maybe(
            5 / 3799 * 5, f"{random.randint(40, 100)}: submitted"
        ),
        "S2 Status: C0 S1's Gen: Drought": maybe(
            3 / 3799 * 5, f"{random.randint(40, 180)}: submitted"
        ),
        "S2 Status: C0 S1's Gen: Acid Soil": maybe(
            9 / 3799 * 5, f"{random.randint(40, 160)} *"
        ),
        "S2 Status: C0 S1's Gen: Waterlo": maybe(
            2 / 3799 * 5, f"{random.randint(100, 130)}: submitted"
        ),
        "S1 Status: C0 S1's Gen: Fusarium": maybe(1 / 3799 * 5, "152: *"),
        "S1 Status: C0 S1's Gen: High Lysine": maybe(
            2 / 3799 * 5, wc(["50+", "High lysine"], [1, 1])
        ),
        "APN2": apn if has_disease else None,
        "Bt all in data": bt_val,
        "GT all in": gt_val,
        "Chem Prof RR test": maybe(4 / 3799 * 3, "Negative"),
        "BtGt Positive/Remarks": btgt_summary,
    }
    return rec


# ══════════════════════════════════════════════════════════════════
# 11. GENERATE + WRITE
# ══════════════════════════════════════════════════════════════════


def generate_dataset(n=500):
    records = [generate_record(i + 1) for i in range(n)]
    return pd.DataFrame(records)


def write_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Basic_Information"

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=9)
    alt = PatternFill("solid", fgColor="EBF3FB")
    plain = PatternFill("solid", fgColor="FFFFFF")

    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = alt if ri % 2 == 0 else plain
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    ws.freeze_panes = "A2"
    for ci, col_name in enumerate(df.columns, 1):
        col_vals = df.iloc[:, ci - 1].astype(str).str.len()
        max_len = max(
            len(str(col_name)),
            int(col_vals.dropna().max()) if col_vals.dropna().size > 0 else 10,
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)
    ws.row_dimensions[1].height = 50

    wb.save(path)
    print(f"Saved: {path}")


def write_csv(df, path):
    df.to_csv(path, index=False)
    print(f"Saved: {path}")


if __name__ == "__main__":
    print(f"Generating {N} synthetic records...")
    df = generate_dataset(N)

    print(f"\nShape: {df.shape}")
    print(f"Columns match original 160? {len(df.columns)} columns")

    # Sanity checks
    ph = pd.to_numeric(df["Plant Height (cm)"], errors="coerce").dropna()
    eh = pd.to_numeric(df["Ear Height (cm)"], errors="coerce").dropna()
    print(f"\nPlant Height: {ph.min():.1f}–{ph.max():.1f}, mean={ph.mean():.1f}")
    print(f"Ear Height:   {eh.min():.1f}–{eh.max():.1f}, mean={eh.mean():.1f}")

    idx = df["Plant Height (cm)"].notna() & df["Ear Height (cm)"].notna()
    bad = (df.loc[idx, "Ear Height (cm)"] >= df.loc[idx, "Plant Height (cm)"]).sum()
    print(f"EH >= PH violations: {bad}  (expected 0)")

    # anthesis/silking
    anth = pd.to_numeric(df["Time of anthesis (DAP)"], errors="coerce").dropna()
    silk = pd.to_numeric(df["Time of silk Emergence (DAP)"], errors="coerce").dropna()
    idx2 = (
        df["Time of anthesis (DAP)"].notna()
        & df["Time of silk Emergence (DAP)"].notna()
    )
    bad2 = (
        df.loc[idx2, "Time of silk Emergence (DAP)"]
        < df.loc[idx2, "Time of anthesis (DAP)"]
    ).sum()
    print(f"Silk before anthesis violations: {bad2}  (expected 0)")

    print(f"\nKernel Type dist:\n{df['Kernel Type'].value_counts().head(6)}")
    print(f"\nKernel Color (top 8):\n{df['Kernel Color'].value_counts().head(8)}")
    print(f"\nALEURONE COLOR:\n{df['ALEURONE COLOR'].value_counts()}")
    print(f"\nDowny Mildew:\n{df['Downy Mildew'].value_counts()}")
    print(
        f"\nBtGt Positive/Remarks (top 8):\n{df['BtGt Positive/Remarks'].value_counts().head(8)}"
    )

    # Glutinous → white check
    glut = df[df["Kernel Type"] == "glutinous"]
    pct = (glut["Kernel Color"].str.lower().str.startswith("white")).mean() * 100
    print(f"\nGlutinous → white-family kernel: {pct:.1f}%  (expected ~97.5%)")

    # Amylose + Amylopectin ≈ Starch
    has_chem = df[["Starch", "Amylose", "Amylopectin"]].notna().all(axis=1)
    if has_chem.sum() > 0:
        sub = df[has_chem]
        diff = abs(sub["Starch"] - (sub["Amylose"] + sub["Amylopectin"]))
        print(
            f"Starch ≈ Amylose+Amylopectin  mean_diff={diff.mean():.2f}  max_diff={diff.max():.2f}"
        )

    out = "synthetic_data.xlsx"
    write_excel(df, out)
    print(f"\nDone. {N} records × {len(df.columns)} columns → {out}")

    csv_out = "synthetic_data.csv"
    write_csv(df, csv_out)
    print(f"Done. {N} records × {len(df.columns)} columns → {csv_out}")
